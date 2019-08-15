import openpyxl
# LENDO DOCUMENTOS EXCEL
wb = openpyxl.load_workbook('ATIVIDADES NUTS 2018.xlsx')
print(type(wb))
# print(wb.get_sheet_names()) # Método depreciado, ele retorna as sheets da planilha, ou seja, as varias planilhas que estão no arquivo

listaDePlanilhas = wb.sheetnames
print(listaDePlanilhas)
# planilhaJaneiro = wb.get_sheet_by_name('JANEIRO') # MÉTODO DEPRECIADO

planilhaJaneiro = wb[listaDePlanilhas[1]] # wb[sheetname]
print(planilhaJaneiro)
print(planilhaJaneiro['A1'])
print(planilhaJaneiro['A1'].value)
print(planilhaJaneiro['B1'].value)
print(planilhaJaneiro['C1'].value)
print(planilhaJaneiro['D1'].value)
print(planilhaJaneiro['E1'].value)
print(planilhaJaneiro['F1'].value)

print(planilhaJaneiro.cell(row=1,column=2))
print(planilhaJaneiro.cell(row=3,column=2).value)

for i in range(1, 22, 1):
    print(i, planilhaJaneiro.cell(row=i,column=2).value)

print("######################################################################")
for i in range(1, 22, 1):
    print(i, planilhaJaneiro.cell(row=i,column=3).value)

print("######################################################################")
for i in range(1, 22, 1):
    print(i, planilhaJaneiro.cell(row=i,column=4).value)

print("######################################################################")
for i in range(1, 22, 1):
    print(i, planilhaJaneiro.cell(row=i,column=5).value)

print("######################################################################")
for i in range(1, 22, 1):
    print(i, planilhaJaneiro.cell(row=i,column=6).value)

""" Podemos determinar o tamanho da planilha usando os métodos get_highest_row() e get_highest_column() do objeto Worksheet """

# tamanhoPlanilha = planilhaJaneiro.max_row_property()
# print(tamanhoPlanilha)

""" Para fazer a conversão de letras para números, chame a função openpyxl.cell.column_index_from_string(). Para converter de números para
letras, chame a função openpyxl.cell.get_column_letter() """

for linhaCelula in planilhaJaneiro['A1':'F21']:
    for celula in linhaCelula:
        print(celula.coordinate, celula.value)
    print("\n")
    print("FIM DA LINHA")


tuplasValores = (planilhaJaneiro['A1':'F21'])
print(tuplasValores)

print(planilhaJaneiro.cell(row=1,column=1))
print(planilhaJaneiro.cell(row=1,column=1).value)

for i in range(1, 22, 1):
    print(i, planilhaJaneiro.cell(row=i,column=1).value) # assim eu percorro toda a coluna da coluna q esta na posição 1