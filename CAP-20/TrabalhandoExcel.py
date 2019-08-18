"""TRABALHANDO COM PLANILHAS EXCEL - JÁ TRABALHEI COM ESSE MODULO.. VOU REVISAR"""
# Documentos Excel = workbook, cada workbook contém varias planilhas = sheets = worksheet
# planilha ativa = active sheet
# cell = celula = linha + coluna

# Instalando o módulo openpyxl
import openpyxl

# Lendo documentos Excel
# Abrindo documentos Excel com o OpenPyXL

planilhaExcel = openpyxl.load_workbook('ATIVIDADES NUTS 2018.xlsx')
print(type(planilhaExcel))

# Obtendo as planilhas do workbook
# lista das planilhas contidas no arquivo excel
listaDePlanilhasNoArquivoExcel = planilhaExcel.sheetnames
print(listaDePlanilhasNoArquivoExcel)

planilhaOutubro = planilhaExcel['OUTUBRO']
print(planilhaOutubro)
print(planilhaOutubro.title)

#Obtendo as células das planilhas
print(planilhaOutubro['A1':'F1']) # Imprimindo os objetos neste intervalo, percorrendo apenas a linha inicial
nomeCelulaA1 = planilhaOutubro['A1'].value
nomeCelulaB1 = planilhaOutubro['B1'].value
nomeCelulaC1 = planilhaOutubro['C1'].value
nomeCelulaD1 = planilhaOutubro['D1'].value
nomeCelulaE1 = planilhaOutubro['E1'].value
nomeCelulaF1 = planilhaOutubro['F1'].value
listaPrimeiraLinhaDaPlanilhaOutubro = []
listaPrimeiraLinhaDaPlanilhaOutubro.append(nomeCelulaA1)
listaPrimeiraLinhaDaPlanilhaOutubro.append(nomeCelulaB1)
listaPrimeiraLinhaDaPlanilhaOutubro.append(nomeCelulaC1)
listaPrimeiraLinhaDaPlanilhaOutubro.append(nomeCelulaD1)
listaPrimeiraLinhaDaPlanilhaOutubro.append(nomeCelulaE1)
listaPrimeiraLinhaDaPlanilhaOutubro.append(nomeCelulaF1)

print(listaPrimeiraLinhaDaPlanilhaOutubro)

## Os objetos Cell, tem atributos row, column, value e coordinate
print(planilhaOutubro['E1'].row) # 1
print(planilhaOutubro['E1'].column) # 5
print(planilhaOutubro['E1'].coordinate) # E1

valorLinhaColuna = planilhaOutubro.cell(row = 3, column = 2)
print(valorLinhaColuna) ## objeto cell
print(valorLinhaColuna.value) ## SIG SIMULAÇÃO EM SAÚDE

for linha in range(1,36,1):
    for coluna in range(1,7,1):
        print(str(linha) +" "+ str(coluna) + " ", planilhaOutubro.cell(row=linha, column=coluna).value)
        #print(linha, planilhaOutubro.cell(row = linha, column = 1).value)

# Fazendo a conversão entre letras e números das colunas
# Obtendo linhas e colunas das planilhas

for linhaCelulaObjeto in planilhaOutubro['A1':'F15']:  # percorre quase toda planilha
    for celulaObjeto in linhaCelulaObjeto:
        print(celulaObjeto.coordinate, " ", celulaObjeto.value)

    print("FIM DA LINHA")


# percorrendo a coluna toda
print("###############################################################################################################")
colunaCompletaB = planilhaOutubro['B'] #isso vira uma tupla
rangeDaColunaC = planilhaOutubro['C'] #isso vira uma tupla
rangeDeColunasDeE = planilhaOutubro['D:E'] #isso vira uma tupla

for j in range(1,35,1):
    print(colunaCompletaB[j].value)

for j in range(1,35,1):
    print(rangeDaColunaC[j].value)

print("###############################################################################################################")
linhaDEZ = planilhaOutubro['11'] # linha completa
for i in range(0,6,1):          # ainda não descobri como pegar o tamanho da tupla
    print(linhaDEZ[i].value)
print("###############################################################################################################")

"""You can also use the Worksheet.iter_rows() method:

>>> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
...    for cell in row:
...        print(cell)
<Cell Sheet1.A1>
<Cell Sheet1.B1>
<Cell Sheet1.C1>
<Cell Sheet1.A2>
<Cell Sheet1.B2>
<Cell Sheet1.C2>

Likewise the Worksheet.iter_cols() method will return columns:

>>> for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
...     for cell in col:
...         print(cell)
<Cell Sheet1.A1>
<Cell Sheet1.A2>
<Cell Sheet1.B1>
<Cell Sheet1.B2>
<Cell Sheet1.C1>
<Cell Sheet1.C2>

"""

print(tuple(planilhaOutubro.rows))
print(tuple(planilhaOutubro.columns))

for linha in planilhaOutubro.values:  # Varrendo toda a planilha
    for valor in linha:
        print(valor)